"""
phase5_postprocessor.py
------------------------
Phase 5 of the Veridooh BFK pipeline.

Takes the merged BKF from Phase 4.5 and produces the final,
upload-ready Excel file.

Responsibilities
----------------
1. Sort rows by Creative Start Date (ASC), then Creative End Date (ASC).
2. Group rows into weekly burst buckets and apply alternating row colours
   so the booker can visually scan week boundaries at a glance.
3. Run final compliance checks (mandatory columns, SOV sum, date range sense).
4. Write a styled .xlsx ready to be uploaded to Google Sheets.

Usage
-----
    from phase5_postprocessor import run_phase5

    report = run_phase5(
        merged_csv   = "campaign_phase45.csv",
        out_xlsx     = "campaign_FINAL.xlsx",
    )
"""

from __future__ import annotations

import warnings
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd

# openpyxl is required for styled Excel output
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    warnings.warn(
        "openpyxl not installed. Styled Excel output will be skipped. "
        "Run: pip install openpyxl"
    )


# ---------------------------------------------------------------------------
# Column references (must match bkf-workflow SKILL.md column headers)
# ---------------------------------------------------------------------------

COL_PANEL_ID   = "Panel ID (or Player ID or Site ID)"
COL_FORMAT     = "Format"
COL_SCREEN_SZ  = "Screen size/ resolution"
COL_CREATIVE   = "Creative File Name"
COL_AD_LEN     = "Ad Length"
COL_SOV        = "SOV (including any vendor content)"
COL_PLAY_INSTR = "Play Instructions"
COL_START      = "Creative Start Date"
COL_END        = "Creative End Date"
COL_COMMENTS   = "Other Comments"

# All mandatory columns per BKF template
MANDATORY_COLS = [
    COL_PANEL_ID,
    "Location Display or Name",
    "Location Identifier (or Site Description or Address)",
    "No. of Screens per media player",
    COL_FORMAT,
    COL_CREATIVE,
    COL_AD_LEN,
    COL_SOV,
    COL_PLAY_INSTR,
    COL_START,
    COL_END,
]

BLANK_TOKENS = {"", "nan", "none", "n/a", "tbc"}

# Week colouring: alternating fills for burst weeks
FILL_WEEK_A = "DCF0DC"   # pale green
FILL_WEEK_B = "FFF2CC"   # pale amber
FILL_HEADER = "2E4057"   # dark blue-grey for the header row
FONT_HEADER = "FFFFFF"


# ---------------------------------------------------------------------------
# Compliance check report
# ---------------------------------------------------------------------------

@dataclass
class ComplianceIssue:
    row_num: int
    panel_id: str
    field: str
    severity: str   # "ERROR" | "WARNING"
    message: str


@dataclass
class PostprocessReport:
    total_rows_in: int          = 0
    total_rows_out: int         = 0
    burst_weeks_detected: int   = 0
    compliance_errors: int      = 0
    compliance_warnings: int    = 0
    issues: list[ComplianceIssue] = field(default_factory=list)

    def summary(self) -> str:
        lines = [
            "=" * 60,
            "Phase 5 — Post-processor Report",
            "=" * 60,
            f"  Rows in                : {self.total_rows_in}",
            f"  Rows out               : {self.total_rows_out}",
            f"  Burst weeks detected   : {self.burst_weeks_detected}",
            f"  Compliance errors      : {self.compliance_errors}",
            f"  Compliance warnings    : {self.compliance_warnings}",
            "=" * 60,
        ]
        if self.issues:
            lines.append("\nCompliance issues (fix before upload):")
            for issue in sorted(self.issues, key=lambda x: (x.severity, x.row_num)):
                lines.append(
                    f"  [{issue.severity}] Row {issue.row_num} "
                    f"(Panel {issue.panel_id!r}) — {issue.field}: {issue.message}"
                )
        return "\n".join(lines)


# ---------------------------------------------------------------------------
# Core post-processor
# ---------------------------------------------------------------------------

class PostProcessor:
    """
    Finalises the merged BKF for upload.

    Parameters
    ----------
    week_start_day : int  ISO weekday for the first day of a burst week.
                         0 = Monday (default), 6 = Sunday.
    """

    def __init__(self, week_start_day: int = 0):
        self.week_start_day = week_start_day

    def process(self, df: pd.DataFrame) -> tuple[pd.DataFrame, PostprocessReport]:
        """
        Sort, label burst weeks, and run compliance checks.

        Returns (processed_df, report).
        The dataframe gains a hidden '__burst_week__' column used for colouring;
        strip it before sharing if you export to CSV.
        """
        df = df.copy()
        report = PostprocessReport(total_rows_in=len(df))

        # 1. Parse and sort by dates
        df = self._parse_dates(df)
        df = self._sort_by_dates(df)

        # 2. Assign burst week labels
        df, report.burst_weeks_detected = self._assign_burst_weeks(df)

        # 3. Compliance checks
        issues = self._check_compliance(df)
        report.issues = issues
        report.compliance_errors   = sum(1 for i in issues if i.severity == "ERROR")
        report.compliance_warnings = sum(1 for i in issues if i.severity == "WARNING")

        report.total_rows_out = len(df)
        return df, report

    # ------------------------------------------------------------------
    # Date parsing & sorting
    # ------------------------------------------------------------------

    def _parse_dates(self, df: pd.DataFrame) -> pd.DataFrame:
        """Parse Creative Start Date and Creative End Date columns to datetime."""
        for col in (COL_START, COL_END):
            if col in df.columns:
                df[f"__{col}__"] = pd.to_datetime(
                    df[col], errors="coerce", dayfirst=True
                )
        return df

    def _sort_by_dates(self, df: pd.DataFrame) -> pd.DataFrame:
        """Sort ascending by start date, then end date, then Panel ID."""
        sort_keys = []
        if f"__{COL_START}__" in df.columns:
            sort_keys.append(f"__{COL_START}__")
        if f"__{COL_END}__" in df.columns:
            sort_keys.append(f"__{COL_END}__")
        if COL_PANEL_ID in df.columns:
            sort_keys.append(COL_PANEL_ID)

        if sort_keys:
            df = df.sort_values(by=sort_keys, ascending=True, na_position="last")
        return df.reset_index(drop=True)

    # ------------------------------------------------------------------
    # Burst week assignment
    # ------------------------------------------------------------------

    def _assign_burst_weeks(self, df: pd.DataFrame) -> tuple[pd.DataFrame, int]:
        """
        Assign a burst week number to each row based on the Creative Start Date.

        A 'burst week' is a Monday–Sunday window (or whichever week_start_day
        you set). All rows whose start date falls in the same calendar week
        are grouped together and will receive the same highlight colour.

        Returns (df_with_week_col, num_distinct_weeks).
        """
        start_col = f"__{COL_START}__"
        if start_col not in df.columns:
            df["__burst_week__"] = 0
            return df, 0

        def _week_key(dt) -> Optional[str]:
            if pd.isna(dt):
                return None
            # ISO year-week gives stable week identifiers
            return dt.strftime("%G-W%V")

        df["__burst_week_key__"] = df[start_col].apply(_week_key)

        # Map week keys to sequential integers (1-based)
        unique_weeks = [
            w for w in df["__burst_week_key__"].unique() if pd.notna(w) and w is not None
        ]
        # Sort chronologically
        unique_weeks.sort()
        week_to_num = {w: i + 1 for i, w in enumerate(unique_weeks)}

        df["__burst_week__"] = df["__burst_week_key__"].map(
            lambda w: week_to_num.get(w, 0)
        )
        df = df.drop(columns=["__burst_week_key__"])

        return df, len(unique_weeks)

    # ------------------------------------------------------------------
    # Compliance checks
    # ------------------------------------------------------------------

    def _check_compliance(self, df: pd.DataFrame) -> list[ComplianceIssue]:
        issues: list[ComplianceIssue] = []

        for idx, row in df.iterrows():
            row_num  = int(idx) + 2    # +2: 1-indexed + header row in Excel
            panel_id = str(row.get(COL_PANEL_ID, "")).strip()

            # Check mandatory fields
            for col in MANDATORY_COLS:
                if col not in df.columns:
                    continue
                val = str(row.get(col, "")).strip().lower()
                if val in BLANK_TOKENS:
                    # Creative File Name is allowed to be blank for partial fill
                    severity = "WARNING" if col == COL_CREATIVE else "ERROR"
                    issues.append(ComplianceIssue(
                        row_num  = row_num,
                        panel_id = panel_id,
                        field    = col,
                        severity = severity,
                        message  = f"Mandatory field is blank or missing",
                    ))

            # Check SOV
            sov_val = str(row.get(COL_SOV, "")).strip()
            if sov_val and sov_val.lower() not in BLANK_TOKENS:
                try:
                    sov = float(sov_val)
                    if not (0 < sov <= 100):
                        issues.append(ComplianceIssue(
                            row_num  = row_num,
                            panel_id = panel_id,
                            field    = COL_SOV,
                            severity = "ERROR",
                            message  = f"SOV value {sov} is outside valid range (0–100)",
                        ))
                except ValueError:
                    issues.append(ComplianceIssue(
                        row_num  = row_num,
                        panel_id = panel_id,
                        field    = COL_SOV,
                        severity = "ERROR",
                        message  = f"SOV value {sov_val!r} is not numeric",
                    ))

            # Check date range sense (start must be before or equal to end)
            start_dt = row.get(f"__{COL_START}__")
            end_dt   = row.get(f"__{COL_END}__")
            if pd.notna(start_dt) and pd.notna(end_dt):
                if start_dt > end_dt:
                    issues.append(ComplianceIssue(
                        row_num  = row_num,
                        panel_id = panel_id,
                        field    = f"{COL_START} / {COL_END}",
                        severity = "ERROR",
                        message  = (
                            f"Start date {start_dt.date()} is after "
                            f"end date {end_dt.date()}"
                        ),
                    ))

            # Play instructions must not be "n/a"
            play = str(row.get(COL_PLAY_INSTR, "")).strip().lower()
            if play == "n/a":
                issues.append(ComplianceIssue(
                    row_num  = row_num,
                    panel_id = panel_id,
                    field    = COL_PLAY_INSTR,
                    severity = "ERROR",
                    message  = 'Play Instructions is "n/a" — BKF rules require "TBC" minimum',
                ))

        return issues


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

class BFKExcelWriter:
    """
    Writes the final BKF dataframe to a styled .xlsx file.
    Applies alternating burst-week row colours and formats the header.
    """

    FILLS = {
        1: PatternFill("solid", fgColor=FILL_WEEK_A),
        0: PatternFill("solid", fgColor=FILL_WEEK_B),  # default / odd-indexed weeks
    }

    def write(self, df: pd.DataFrame, out_path: Path) -> None:
        if not OPENPYXL_AVAILABLE:
            # Fallback: write plain CSV
            csv_path = out_path.with_suffix(".csv")
            _clean_export(df).to_csv(csv_path, index=False)
            print(f"[Warning] openpyxl not available. Plain CSV written to: {csv_path}")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "VMO"

        export_df = _clean_export(df)
        cols      = list(export_df.columns)

        # Write header (row 1)
        header_fill = PatternFill("solid", fgColor=FILL_HEADER)
        header_font = Font(color=FONT_HEADER, bold=True, size=10)
        for col_idx, col_name in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill   = header_fill
            cell.font   = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Write data rows with burst-week colouring
        week_col = "__burst_week__"
        for row_idx, (_, data_row) in enumerate(df.iterrows(), start=2):
            burst_week = int(data_row.get(week_col, 0))
            fill = PatternFill("solid", fgColor=FILL_WEEK_A if burst_week % 2 == 1 else FILL_WEEK_B)

            for col_idx, col_name in enumerate(cols, start=1):
                val  = data_row.get(col_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if pd.notna(val) else "")
                cell.fill      = fill
                cell.font      = Font(size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Auto-fit column widths (capped at 40 chars)
        for col_idx, col_name in enumerate(cols, start=1):
            col_letter = get_column_letter(col_idx)
            max_len = max(
                len(col_name),
                *(
                    len(str(ws.cell(row=r, column=col_idx).value or ""))
                    for r in range(2, ws.max_row + 1)
                ),
            )
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Freeze the header row
        ws.freeze_panes = "A2"

        wb.save(out_path)
        print(f"Styled Excel written to: {out_path}")


def _clean_export(df: pd.DataFrame) -> pd.DataFrame:
    """Remove internal helper columns before writing."""
    drop = [c for c in df.columns if c.startswith("__")]
    return df.drop(columns=drop, errors="ignore")


# ---------------------------------------------------------------------------
# Convenience entry point
# ---------------------------------------------------------------------------

def run_phase5(
    merged_csv: str | Path,
    out_xlsx: Optional[str | Path] = None,
    week_start_day: int = 0,
) -> PostprocessReport:
    """
    Load Phase 4.5 merged BKF CSV, post-process, and write final styled Excel.

    Parameters
    ----------
    merged_csv     : path to phase45 merged BKF CSV
    out_xlsx       : output .xlsx path (default: <stem>_FINAL.xlsx)
    week_start_day : 0=Monday, 6=Sunday for burst week grouping
    """
    csv_path = Path(merged_csv)
    df = pd.read_csv(csv_path, dtype=str).fillna("")

    processor = PostProcessor(week_start_day=week_start_day)
    df_out, report = processor.process(df)

    print(report.summary())

    if report.compliance_errors > 0:
        print(
            f"\n⚠  {report.compliance_errors} compliance ERROR(s) found. "
            "Fix these before uploading to Google Sheets."
        )

    # Write Excel
    if out_xlsx is None:
        out_xlsx = csv_path.with_name(
            csv_path.stem.replace("_phase45", "") + "_FINAL.xlsx"
        )
    writer = BFKExcelWriter()
    writer.write(df_out, Path(out_xlsx))

    return report


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Phase 5 Post-processor")
    parser.add_argument("merged_csv", help="Phase 4.5 merged BKF CSV")
    parser.add_argument("--out",      help="Output .xlsx path")
    parser.add_argument(
        "--week-start", type=int, default=0,
        help="ISO weekday for burst week start: 0=Mon (default), 6=Sun"
    )
    args = parser.parse_args()

    run_phase5(
        merged_csv     = args.merged_csv,
        out_xlsx       = args.out,
        week_start_day = args.week_start,
    )
