"""
Read-only inspector for BKF/IO/MI Excel files.
Dumps sheet names, locates the header row (row containing 'Panel ID' or similar),
and prints header + sample data rows WITH per-cell data_type and number_format,
so we can see whether dates are real date cells or text, and how SOV is stored.
"""
import sys
from openpyxl import load_workbook


def find_header_row(ws, max_scan=30):
    """Return (row_idx, [cell values]) for the first row that looks like a header."""
    needles = ("panel id", "location", "format", "creative", "sov", "play")
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        text = " ".join(str(v).lower() for v in vals if v is not None)
        hits = sum(1 for n in needles if n in text)
        if hits >= 2:
            return r, vals
    # fallback: first non-empty row
    for r in range(1, min(ws.max_row, max_scan) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in vals):
            return r, vals
    return 1, []


def inspect(path):
    print("=" * 100)
    print(f"FILE: {path}")
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        print(f"  !! could not open: {e}")
        return
    print(f"  sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    print(f"  first sheet '{ws.title}': dims={ws.max_row} rows x {ws.max_column} cols")

    hdr_row, _ = find_header_row(ws)
    print(f"  header row detected at Excel row {hdr_row}")
    cols = [(c, ws.cell(row=hdr_row, column=c).value) for c in range(1, ws.max_column + 1)]
    print("  COLUMNS:")
    for c, v in cols:
        if v is not None and str(v).strip():
            print(f"    [{c:2}] {v!r}")

    print("  SAMPLE DATA (value | data_type | number_format):")
    shown = 0
    r = hdr_row + 1
    while r <= ws.max_row and shown < 3:
        row_has_data = any(
            ws.cell(row=r, column=c).value not in (None, "")
            for c in range(1, ws.max_column + 1)
        )
        if row_has_data:
            print(f"    --- Excel row {r} ---")
            for c, hv in cols:
                if hv is None or not str(hv).strip():
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.value in (None, ""):
                    continue
                print(f"      [{c:2}] {str(hv)[:38]:38} = {str(cell.value)[:34]!r:36} "
                      f"| {cell.data_type} | {cell.number_format}")
            shown += 1
        r += 1
    wb.close()


if __name__ == "__main__":
    for p in sys.argv[1:]:
        inspect(p)
