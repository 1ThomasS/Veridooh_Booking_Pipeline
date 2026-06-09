"""
Targeted BKF data inspector. Assumes the real header is at a known Excel row
(default 14 for the Veridooh 'Supplier wc' template) and prints the first N
populated DATA rows below it, with per-cell value | data_type | number_format.
Also reports how many data rows are populated (Panel ID non-empty).
"""
import sys
from openpyxl import load_workbook

HEADER_ROW = 14
N_SAMPLE = 4


def inspect(path):
    print("=" * 100)
    print(f"FILE: {path.split('/')[-1]}")
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    hdr = [(c, ws.cell(row=HEADER_ROW, column=c).value) for c in range(1, 17)]

    # count populated data rows (col 1 Panel ID non-empty)
    n_pop = 0
    last_row = HEADER_ROW
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            n_pop += 1
            last_row = r
    print(f"  populated data rows (Panel ID non-empty): {n_pop} (rows {HEADER_ROW+1}..{last_row})")

    shown = 0
    r = HEADER_ROW + 1
    while r <= ws.max_row and shown < N_SAMPLE:
        if ws.cell(row=r, column=1).value not in (None, ""):
            print(f"  --- data row {r} ---")
            for c, hv in hdr:
                cell = ws.cell(row=r, column=c)
                if cell.value in (None, "") and (hv is None or not str(hv).strip()):
                    continue
                label = (str(hv)[:34] if hv else f"col{c}")
                print(f"    [{c:2}] {label:34} = {str(cell.value)[:32]!r:34} "
                      f"| {cell.data_type} | {cell.number_format}")
            shown += 1
        r += 1
    wb.close()


if __name__ == "__main__":
    for p in sys.argv[1:]:
        inspect(p)
