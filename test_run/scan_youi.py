"""
Aggregate scanner over filled Youi 'Booking Forms' to learn the PASSING format:
- header row, sheet name pattern
- per-column fill rate
- distinct Play Instructions values
- distinct SOV values + their number_format / data_type
- date column data_type + number_format distribution
- Panel ID samples
Read-only; prints a consolidated report.
"""
import sys, glob, os
from collections import Counter, defaultdict
from openpyxl import load_workbook
import warnings
warnings.simplefilter("ignore")

COLS = {
    1: "PanelID", 2: "PanelDesc", 3: "LocDisplay", 4: "LocIdentifier",
    5: "NoScreens", 6: "ActiveDisplay", 7: "Format", 8: "ScreenSize",
    9: "Creative", 10: "AdLength", 11: "SOV", 12: "PlayInstr",
    13: "StartDate", 14: "EndDate", 15: "Times", 16: "OtherComments",
}


def find_header(ws):
    for r in range(1, min(ws.max_row, 40) + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith("panel id"):
            return r
    return None


def main(root):
    files = sorted(glob.glob(os.path.join(root, "**", "Booking Forms", "*.xlsx"), recursive=True))
    files = [f for f in files if "~$" not in f]
    print(f"Found {len(files)} booking-form files under {root}\n")

    fill = defaultdict(int)          # col -> #rows with non-empty value
    total_rows = 0
    play_vals = Counter()
    sov_vals = Counter()
    sov_fmt = Counter()
    date_type = Counter()
    date_fmt = Counter()
    panelid_samples = []
    noscreens_examples = []
    locid_blank_files = []
    sheet_names = []
    files_scanned = 0
    files_empty = 0

    for f in files:
        try:
            wb = load_workbook(f, data_only=False)
        except Exception:
            continue
        ws = wb[wb.sheetnames[0]]
        hr = find_header(ws)
        if hr is None:
            wb.close(); continue
        sheet_names.append(ws.title)
        n_here = 0
        locid_blank_here = 0
        for r in range(hr + 1, ws.max_row + 1):
            pid = ws.cell(row=r, column=1).value
            if pid in (None, ""):
                continue
            n_here += 1; total_rows += 1
            if len(panelid_samples) < 40:
                panelid_samples.append(str(pid))
            for c in COLS:
                v = ws.cell(row=r, column=c).value
                if v not in (None, ""):
                    fill[c] += 1
            # play
            pv = ws.cell(row=r, column=12).value
            if pv not in (None, ""):
                play_vals[str(pv)] += 1
            # sov
            sv = ws.cell(row=r, column=11)
            if sv.value not in (None, ""):
                sov_vals[str(sv.value)] += 1
                sov_fmt[(sv.data_type, sv.number_format)] += 1
            # dates
            for dc in (13, 14):
                dcell = ws.cell(row=r, column=dc)
                if dcell.value not in (None, ""):
                    date_type[dcell.data_type] += 1
                    date_fmt[str(dcell.number_format)] += 1
            # noscreens
            ns = ws.cell(row=r, column=5).value
            if ns not in (None, "") and len(noscreens_examples) < 10:
                noscreens_examples.append((os.path.basename(f)[:30], str(ns)))
            # locid blank
            if ws.cell(row=r, column=4).value in (None, ""):
                locid_blank_here += 1
        if n_here == 0:
            files_empty += 1
        else:
            files_scanned += 1
        if locid_blank_here:
            locid_blank_files.append((os.path.basename(f)[:50], locid_blank_here, n_here))
        wb.close()

    print(f"Files with data: {files_scanned} | empty/template: {files_empty} | total data rows: {total_rows}\n")
    print("PER-COLUMN FILL RATE (% of data rows non-empty):")
    for c, name in COLS.items():
        pct = 100 * fill[c] / total_rows if total_rows else 0
        print(f"  [{c:2}] {name:14} {fill[c]:5}/{total_rows}  {pct:5.1f}%")

    print("\nDISTINCT Play Instructions (top 25):")
    for v, n in play_vals.most_common(25):
        print(f"  {n:5} x  {v!r}")

    print("\nSOV cell (data_type, number_format) counts:")
    for k, n in sov_fmt.most_common():
        print(f"  {n:5} x  {k}")
    print("DISTINCT SOV values (top 20):")
    for v, n in sov_vals.most_common(20):
        print(f"  {n:5} x  {v!r}")

    print("\nDATE data_type counts:", dict(date_type))
    print("DATE number_format counts:", dict(date_fmt))

    print("\nPanel ID samples:", panelid_samples[:40])
    print("\nNo. of Screens non-empty examples:", noscreens_examples)
    print("\nFiles where Location Identifier blank on some rows:")
    for fn, b, n in locid_blank_files[:20]:
        print(f"  {fn}: {b}/{n} rows blank")
    print("\nSheet name samples:", sheet_names[:15])


if __name__ == "__main__":
    main(sys.argv[1])
