"""
BKF data inspector that auto-detects the real header row (col-1 cell starts with
'Panel ID') and prints the first N populated data rows with value|data_type|number_format.
"""
import sys
from openpyxl import load_workbook

N_SAMPLE = 4


def find_header(ws):
    for r in range(1, min(ws.max_row, 40) + 1):
        v = ws.cell(row=r, column=1).value
        if v and str(v).strip().lower().startswith("panel id"):
            return r
    return None


def inspect(path):
    print("=" * 100)
    print(f"FILE: {path.split('/')[-1]}")
    wb = load_workbook(path, data_only=False)
    print(f"  sheets: {wb.sheetnames}")
    ws = wb[wb.sheetnames[0]]
    hr = find_header(ws)
    if hr is None:
        print("  !! no 'Panel ID' header found in col 1; dumping row1")
        print("   ", [ws.cell(row=1, column=c).value for c in range(1, 17)])
        wb.close()
        return
    print(f"  header at Excel row {hr}")
    hdr = [(c, ws.cell(row=hr, column=c).value) for c in range(1, ws.max_column + 1)]

    n_pop, last = 0, hr
    for r in range(hr + 1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            n_pop += 1
            last = r
    print(f"  populated data rows: {n_pop} (rows {hr+1}..{last})")

    shown, r = 0, hr + 1
    while r <= ws.max_row and shown < N_SAMPLE:
        if ws.cell(row=r, column=1).value not in (None, ""):
            print(f"  --- data row {r} ---")
            for c, hv in hdr:
                if hv is None or not str(hv).strip():
                    continue
                cell = ws.cell(row=r, column=c)
                if cell.value in (None, ""):
                    continue
                print(f"    [{c:2}] {str(hv)[:32]:32} = {str(cell.value)[:30]!r:32} "
                      f"| {cell.data_type} | {cell.number_format}")
            shown += 1
        r += 1
    wb.close()


if __name__ == "__main__":
    for p in sys.argv[1:]:
        inspect(p)
